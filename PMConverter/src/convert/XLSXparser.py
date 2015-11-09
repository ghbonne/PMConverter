__author__ = 'Project management group 8, Ghent University 2015'
__license__ = ["BSD", "MIT/expat"]
__credits__ = ["John McNamara"]

import xlsxwriter
import openpyxl
import re, sys
import datetime
from operator import attrgetter

from objects.activity import Activity
from objects.activitytracking import ActivityTrackingRecord
from objects.baselineschedule import BaselineScheduleRecord
from objects.projectobject import ProjectObject
from objects.resource import Resource, ResourceType
from objects.riskanalysisdistribution import RiskAnalysisDistribution
from objects.trackingperiod import TrackingPeriod
from convert.fileparser import FileParser
from objects.agenda import Agenda
from exceptions import XLSXParseError

# DEBUG
import csv, os


class XLSXParser(FileParser):
    """
    Class to convert ProjectObjects to .xlsx files and vice versa. Shout out to John McNamara for his xlsxwriter library
    and the guys from openpyxl.
    """

    def __init__(self):
        super(XLSXParser, self).__init__()

    """ ---------------------- MAIN FUNCTION TO READ FROM EXCEL ------------------------- """
    def to_schedule_object(self, file_path_input):
        """Reads an Excel file and loads its data in a project object."""

        workbook = openpyxl.load_workbook(file_path_input, data_only=True)
        project_control_sheets = []
        agenda_sheet = None
        for name in workbook.get_sheet_names():
            if "Baseline Schedule" in name:
                activities_sheet = workbook.get_sheet_by_name(name)
            elif "Resources" in name:
                resource_sheet = workbook.get_sheet_by_name(name)
            elif "Risk Analysis" in name:
                risk_analysis_sheet = workbook.get_sheet_by_name(name)
            elif "TP" in name:
                project_control_sheets.append(workbook.get_sheet_by_name(name))
            elif "Agenda" in name:
                agenda_sheet = workbook.get_sheet_by_name(name)

        # Process the agenda sheet first to have an agenda to parse all dates in the other sheets
        if agenda_sheet:
            agenda = self.process_agenda(agenda_sheet)
        else:
            agenda = Agenda()
        # Next we process the resources sheet, we store them in a dict, with index the resource name, to access them
        # easily later when we are processing the activities.

        resources_dict = self.process_resources(resource_sheet)
        # Then we process the risk analysis sheet, again everything is stored in a dict, now index is activity_id.
        risk_analysis_dict = self.process_risk_analysis(risk_analysis_sheet)

        # Finally, the sheet with activities is processed, using the dicts we created above.
        # Again, new dicts are created, to process all tracking periods more easily
        activities_dict, activityGroups_dict, activityGroup_to_childActivities_dict, project_name = self.process_baseline_schedule(activities_sheet, resources_dict, risk_analysis_dict, agenda)
        tracking_periods = self.process_project_controls(project_control_sheets, activities_dict, activityGroups_dict, activityGroup_to_childActivities_dict, agenda)

        return ProjectObject(name= project_name,activities=[i[1] for i in sorted(activities_dict.values(), key=lambda x: x[1].wbs_id)],
                             resources=sorted(resources_dict.values(), key=lambda x: x.resource_id),
                             tracking_periods=tracking_periods, agenda=agenda)

    """ READING PRIVATE FUNCTIONS """
    @staticmethod
    def process_predecessors(predecessors, agenda):
        """
        Reads an Excel Baseline Schedule tab's predecessors field
        :return: list of [(predecessor_activity_id, predecessor_relation, predecessor_lag), ...]
        """
        if predecessors and predecessors.strip():
            activity_predecessors = []
            for predecessor in predecessors.split(";"):
                temp = re.split('\-|\+', predecessor)
                # The last two characters are the relation type (e.g. xFS or xxFS), activity can be variable
                # number of digits
                predecessor_part = temp[0].strip() # remove excess spaces
                predecessor_activity_id = int(predecessor_part[0:-2])
                predecessor_relation = predecessor_part[-2:]
                if len(temp) == 2:
                    # Was it a + or -?
                    minus_plus = predecessor.split("-")
                    if len(minus_plus) == 2:
                        # It was a -
                        predecessor_lag = -agenda.convert_durationString_to_workingHours(temp[1]) #extract workinghours from string
                    else:
                        # It was a +
                        predecessor_lag = agenda.convert_durationString_to_workingHours(temp[1]) #extract workinghours from string
                else:
                    predecessor_lag = 0
                activity_predecessors.append((predecessor_activity_id, predecessor_relation, predecessor_lag))
            return activity_predecessors
        return []

    @staticmethod
    def process_successors(successors, agenda):
        """
        Reads an Excel Baseline Schedule tab's successors field
        :return: list of [(successor_activity_id, successor_relation, successor_lag), ...]
        """
        if successors and successors.strip():
            activity_successors = []
            for successor in successors.split(";"):
                temp = re.split('\-|\+', successor)
                # The first two characters are the relation type (e.g. FSx or FSxx), activity can be variable number of digits
                successor_part = temp[0].strip() # remove excess spaces
                successor_activity_id = int(successor_part[2:])
                successor_relation = successor_part[0:2]
                if len(temp) == 2:
                    minus_plus = successor.split("-")
                    if len(minus_plus) == 2:
                        # It was a -
                        successor_lag = -agenda.convert_durationString_to_workingHours(temp[1]) #extract workinghours from string
                    else:
                        # It was a +
                        successor_lag = agenda.convert_durationString_to_workingHours(temp[1]) #extract workinghours from string
                else:
                    successor_lag = 0
                activity_successors.append((successor_activity_id, successor_relation, successor_lag))
            return activity_successors
        return []

    @staticmethod
    def get_nr_of_header_lines(sheet):
        """
        Counts the number of rows from the top until a digit (ID number) is found in the first column.
        :return: int, number of first row with a digit
        """
        header_lines = 1
        while sheet.cell(row=header_lines, column=1).value is None \
                or (type(sheet.cell(row=header_lines, column=1).value) is not int
                    and not sheet.cell(row=header_lines, column=1).value.isdigit()):
            header_lines += 1
            if header_lines == 100:  # An after-deadline hack to avoid infinite loops
                print("Error:XLSXParser:get_nr_of_header_lines: After-deadline hack to avoid infinite loops was needed!")
                break
        return header_lines

    @staticmethod
    def read_date(date):
        """
        Processes an Excel string to a datetime
        :return: datetime
        """
        date_raw = None
        if type(date) is not datetime.datetime:
            date_raw = datetime.datetime.utcfromtimestamp(((date - 25569)*86400)).date()  # ugly hack to convert
        else:
            date_raw = date

        # round input dates to the hour:
        return datetime.datetime(year= date_raw.year, month= date_raw.month, day= date_raw.day, hour= date_raw.hour if date_raw.minute < 30 else date_raw.hour + 1)

    @staticmethod
    def process_activity_resource_assignments(activity_assignment_str, activity_resources, resources_dict, activity_baseline_duration_workingHours):
        """
        Processes the activity resource assignment field of the Baseline schedule table.
        :return: total cost of resource assignments and implicitly add resource assignments to the activity.resources
        """
        resources_cost = 0.0
        for activity_resource_readString in activity_assignment_str.split(';'):
            if not activity_resource_readString or not activity_resource_readString.strip():
                continue  # if empty string => continue
            activity_resource = resources_dict[(activity_resource_readString.split("[")[0]).strip()]
            activity_resource_demand = 1.0
            activity_resource_demand_fixed = False  # for consumable resources only
            if len(activity_resource_readString.split("[")) > 1:
                activity_resource_demand_str = activity_resource_readString.split("[")[1].split(" ")[0]
                if activity_resource_demand_str[-1] is ']':
                    activity_resource_demand_str = activity_resource_demand_str[:-1]
                activity_resource_demand_fixed = activity_resource_demand_str.endswith("F")
                if activity_resource_demand_fixed:
                    activity_resource_demand_str = activity_resource_demand_str[:-1]
                    # check if resource is consumable:
                    if activity_resource.resource_type != ResourceType.CONSUMABLE:
                        print("XLSXParser:process_baseline_schedule:Found a fixed resource assignment to {0}, which is a non-consumable resource type!".format(activity_resource.name))
                        activity_resource_demand_fixed = False

                activity_resource_demand = float(activity_resource_demand_str.translate(str.maketrans(",", "."))) # demand can be integer or float!

            # calculate cost of resource assignment
            resource_demand_cost = Resource.calculate_resource_assignment_cost(activity_resource, activity_resource_demand, activity_resource_demand_fixed, activity_baseline_duration_workingHours)
            resources_cost += resource_demand_cost
            # update resource its total cost:
            activity_resource.total_resource_cost += resource_demand_cost

            activity_resources.append((activity_resource, activity_resource_demand, activity_resource_demand_fixed))

        return resources_cost

    """ SHEET PROCESSING FUNCTIONS """
    def process_agenda(self, agenda_sheet):
        """
        Reads an Excel Agenda tab
        :return: Agenda object
        """
        working_days = [0]*7
        working_hours = [0]*24
        holidays = []

        if not (agenda_sheet.cell(row=1, column=1).value and agenda_sheet.cell(row=1, column=4).value and agenda_sheet.cell(row=1, column=7).value):
            raise XLSXParseError("Not all required columns are on the required position of the Agenda sheet.")
        if not agenda_sheet.cell(row=25, column=2).value:
            raise XLSXParseError("24 hour fields are required in the Agenda sheet.")
        if not agenda_sheet.cell(row=8, column=5).value:
            raise XLSXParseError("7 days fields are required in the Agenda sheet.")

        # Iterate over all hour fields
        for i in range(0, 24):
            if agenda_sheet.cell(row=i+2, column=2).value.lower() == "yes":
                working_hours[i] = 1

        # Iterate over all days
        for i in range(0, 7):
            if agenda_sheet.cell(row=i+2, column=5).value.lower() == "yes":
                working_days[i] = 1

        # Check if any holidays are given
        i = 2
        while agenda_sheet.cell(row=i, column=7).value:
            holidays.append(self.read_date(agenda_sheet.cell(row=i, column=7).value))
            i += 1

        return Agenda(working_hours=working_hours, working_days=working_days, holidays=holidays)

    def process_project_controls(self, project_control_sheets, activities_dict, activityGroups_dict, activityGroup_to_childActivities_dict, agenda):
        """
        Reads Excel Tracking period tabs
        :paarm project_control_sheets: list of openpyxl worksheets of tracking periods
        :param activities_dict: dict of all activities and activitygroups with 'key: value' = 'activityId: (row, Activity)'
        :param activityGroups_dict: dict of only activitygroups with 'key: value' = 'activityId: Activity'
        :param activityGroup_to_childActivities_dict: dict which links an activitygroupId to its child activities' ID, 'key: value' = 'activityId: [ActivityId1, ActivityId2,...]'
        :param agenda: Agenda object
        :return: list of trackingperiod objects
        """
        tracking_periods = []

        for project_control_sheet in project_control_sheets:
            if project_control_sheet.cell(row=1, column=3).value:
                tp_date = self.read_date(project_control_sheet.cell(row=1, column=3).value)
            else:
                raise XLSXParseError("process_project_controls: No status date found on tracking period sheet: {0}".format(project_control_sheet.title))

            tp_name = project_control_sheet.cell(row=1, column=6).value
            if not tp_name or not str(tp_name).strip():
                # no TP name was given, we can take the title of the sheet as default value
                tp_name = project_control_sheet.title

            act_track_records = []
            currentTrackingPeriod_records_dict = {}

            for curr_row in range(self.get_nr_of_header_lines(project_control_sheet), project_control_sheet.get_highest_row()+1):
                # don't read activityGroups, add them later
                activity_id_field = project_control_sheet.cell(row=curr_row, column=1).value
                if activity_id_field is not None and str(activity_id_field).strip(): # if not None or empty or only spaces
                    # check if activity_id field corresponds with an activityGroup
                    if int(activity_id_field) in activityGroups_dict:
                        # process groups later
                        continue
                    elif int(activity_id_field) not in activities_dict:
                        raise XLSXParseError(("An error occurred while processing tracking period sheet: {0}\n" + \
                                "Unknown activity ID {1} found on row {2}.\n").format(project_control_sheet.title, activity_id_field, curr_row))
                else:
                    # Eliminate empty rows below the actual table => do not throw an exception for this:
                    print("XLSXparser:process_project_controls: Row {0} in TP sheet ({1}) does not contain a valid activity ID: {2}".format(curr_row, project_control_sheet.title, activity_id_field))
                    continue

                # processing an activity tracking record
                activity_id = int(activity_id_field)
                currentActivity = activities_dict[activity_id][1]
                actual_start = datetime.datetime.max  # Set a default value in case there is nothing in that cell
                actual_duration = datetime.timedelta(0) # default
                actual_duration_hours = 0 # default
                if project_control_sheet.cell(row=curr_row, column=12).value is not None and str(project_control_sheet.cell(row=curr_row, column=12).value).strip():
                    actual_start = self.read_date(project_control_sheet.cell(row=curr_row, column=12).value)

                actualDuration_field = project_control_sheet.cell(row=curr_row, column=13).value
                if actualDuration_field is not None and isinstance(actualDuration_field, str) and actualDuration_field.strip():
                    actual_duration_hours = agenda.convert_durationString_to_workingHours(project_control_sheet.cell(row=curr_row, column=13).value)
                    actual_duration = agenda.get_workingDuration_timedelta(actual_duration_hours)

                prc_dev = 0.0
                if project_control_sheet.cell(row=curr_row, column=18).value and str(project_control_sheet.cell(row=curr_row, column=18).value).strip():
                    prc_dev = float(project_control_sheet.cell(row=curr_row, column=18).value)

                actual_cost = 0.0
                if project_control_sheet.cell(row=curr_row, column=19).value and str(project_control_sheet.cell(row=curr_row, column=19).value).strip():
                    actual_cost = float(project_control_sheet.cell(row=curr_row, column=19).value)

                percentage_completed_str = project_control_sheet.cell(row=curr_row, column=21).value  # always given
                if percentage_completed_str is None or not str(percentage_completed_str).strip():
                    raise XLSXParseError("process_project_controls: No percentage completed found on row {0} of sheet {1}".format(curr_row, project_control_sheet.title))

                # In the test file, some percentages are strings, some are ints, some are floats (#YOLO)
                if type(percentage_completed_str) != str:
                    percentage_completed_str *= 100
                percentage_completed_str = str(percentage_completed_str)
                if not percentage_completed_str[-1].isdigit():
                    percentage_completed_str = percentage_completed_str[:-1] # remove the % in the string
                ### percentages can also be smaller than 1!!!!
                #if float(percentage_completed_str) < 1: 
                #    percentage_completed_str = float(float(percentage_completed_str)*100)
                percentage_completed = float(percentage_completed_str)
                if percentage_completed > 100:
                    if percentage_completed > 100 + 1e-5:
                        print("Gilles V has made a stupid error by putting some ifs & transformations in XLSXparser.py "
                            "with the variable percentage_completed_str")
                    # fix percentage completed to 100.0
                    percentage_completed = 100.0

                # convert percentage completed to remaining duration
                if percentage_completed > 1e-5 and actual_duration_hours > 0:
                    remainingDuration_hours = round(actual_duration_hours / float(percentage_completed) * 100 - actual_duration_hours)
                else:
                    remainingDuration_hours = round(currentActivity.baseline_schedule.duration.days * agenda.get_working_hours_in_a_day() + currentActivity.baseline_schedule.duration.seconds / 3600.0)

                remaining_duration = agenda.get_workingDuration_timedelta(remainingDuration_hours)

                # derive other fields:
                actual_cost_calced, earned_value, planned_actual_cost, planned_remaining_cost, planned_value, remaining_cost = \
                        ActivityTrackingRecord.calculate_activityTrackingRecord_derived_values(activities_dict[activity_id][1],0, actual_duration_hours, agenda, percentage_completed, prc_dev,
                                                                                               remainingDuration_hours, tp_date, actual_start)
                pac_dev = actual_cost - planned_actual_cost

                # determine tracking status:
                if actual_start == datetime.datetime.max:
                    tracking_status = "Not Started"
                elif abs(percentage_completed - 100) < 1e-10: # compare float
                    tracking_status = "Finished"
                else:
                    tracking_status = "Started"

                act_track_record = ActivityTrackingRecord(activity=activities_dict[activity_id][1],
                                                          actual_start=actual_start, actual_duration=actual_duration,
                                                          planned_actual_cost=planned_actual_cost, planned_remaining_cost=planned_remaining_cost,
                                                          remaining_duration=remaining_duration, deviation_pac=pac_dev,
                                                          deviation_prc=prc_dev, actual_cost=actual_cost,
                                                          remaining_cost=remaining_cost,
                                                          percentage_completed=percentage_completed,
                                                          tracking_status=tracking_status, earned_value=earned_value,
                                                          planned_value=planned_value)
                act_track_records.append(act_track_record)
                currentTrackingPeriod_records_dict[activity_id] = act_track_record

            # check if activityTrackingRecords found for every activity:
            for act_id in activities_dict.keys():
                # if a low level activity has no tracking record => exception
                if act_id not in activityGroups_dict and act_id not in currentTrackingPeriod_records_dict:
                    raise XLSXParseError(("An error occurred while processing the tracking period sheet: {0}\n" + \
                        "Could not find a row with tracking info about the activity with id:{1}\n").format(project_control_sheet.title, act_id))

            currentTrackingPeriod = TrackingPeriod(tracking_period_name=tp_name, tracking_period_statusdate=tp_date,
                                                   tracking_period_records=act_track_records)
            # calculate aggregated values for activityGroups tracking:
            for activityGroupId in activityGroups_dict.keys():
                childActivityIds = activityGroup_to_childActivities_dict[activityGroupId]
                # construct the activityGroup trackingRecord:
                activityGroup_trackingRecord = ActivityTrackingRecord.construct_activityGroup_trackingRecord(activityGroups_dict[activityGroupId], childActivityIds, currentTrackingPeriod_records_dict, tp_date, agenda)
                currentTrackingPeriod.tracking_period_records.append(activityGroup_trackingRecord)

            # sort currentTrackingPeriod on wbs:
            currentTrackingPeriod.tracking_period_records = sorted(currentTrackingPeriod.tracking_period_records, key=lambda activityTrackingRecord: activityTrackingRecord.activity.wbs_id)

            # add trackingperiod to list
            tracking_periods.append(currentTrackingPeriod)

        # sort tracking periods on status date
        tracking_periods = sorted(tracking_periods, key= lambda x: x.tracking_period_statusdate)
        return tracking_periods

    def process_resources(self, resource_sheet):
        """
        Reads an Excel Resources tab
        :return: dict: the resources in a dict, with as index the resource name
        """
        # We store the resources  in a dict, with as index the resource name, to access them easily later when we
        # are processing the activities.
        resources_dict = {}
        try:
            for curr_row in range(self.get_nr_of_header_lines(resource_sheet), resource_sheet.get_highest_row()+1):
                res_id_str = resource_sheet.cell(row=curr_row, column=1).value
                res_id = -1
                try:
                    # try to cast this string to an int
                    res_id = int(res_id_str)
                except:
                    # This row has no valid ID number => skip this row
                    print("XLSXparser:process_resources: Invalid resource ID at row {0}".format(curr_row))
                    continue

                res_name = str(resource_sheet.cell(row=curr_row, column=2).value).strip() # strip leading and ending spaces
                # check if this resource name is already present:
                if not res_name:
                    raise XLSXParseError(("An error occurred while processing the Excel Resources sheet at row {0} and column B\n" + \
                        "No empty resource names allowed!").format(curr_row))
                elif res_name in resources_dict:
                    raise XLSXParseError(("An error occurred while processing the Excel Resources sheet at row {0} and column B\n" + \
                        "No duplicate resource names allowed: {1} is already defined!").format(curr_row, res_name))

                # parse resource type
                res_type_string = resource_sheet.cell(row=curr_row, column=3).value
                if res_type_string and res_type_string.lower() == ResourceType.CONSUMABLE.value.lower():
                    res_type = ResourceType.CONSUMABLE
                elif res_type_string and res_type_string.lower() == ResourceType.RENEWABLE.value.lower():
                    res_type = ResourceType.RENEWABLE
                else:
                    raise XLSXParseError(("An error occurred while processing the Excel Resources sheet at row {0} and column C\n" + \
                        "Unkown resource type: {1}\n" + \
                        "NOTE: Only 'Renewable' or 'Consumable' are defined.").format(curr_row, res_type_string if res_type_string else "None"))

                if res_type != ResourceType.CONSUMABLE:
                    # Had to cast string -> float -> int (silly Python!)
                    ava_str_split = str(resource_sheet.cell(row=curr_row, column=4).value).split(" ")
                    # Try to read the availability number and unit:
                    res_ava = 0
                    res_unit = ""
                    try:
                        res_ava = int(float(ava_str_split[0].translate(str.maketrans(",", "."))))
                        if len(ava_str_split) > 1:
                            res_unit = ava_str_split[1]
                            # append words back to the unit string if spaces would be present in the unit string
                            for i in range(2, len(ava_str_split)):
                                res_unit += " " + ava_str_split[i]
                        else:
                            res_unit = ""
                    except ValueError:
                        raise XLSXParseError(("An error occurred while processing the Excel Resources sheet at row {0} and column D\n" + \
                                "Can't figure out the availability number and/or resource unit from: {1}\n" + \
                                "NOTE: A space is needed between the resource availability and resource unit.").format(curr_row, resource_sheet.cell(row=curr_row, column=4).value))

                else:
                    res_ava = -1
                    res_unit = ""
                res_cost_use = float(resource_sheet.cell(row=curr_row, column=5).value)
                res_cost_unit = float(resource_sheet.cell(row=curr_row, column=6).value)
                resources_dict[res_name] = Resource(resource_id=res_id, name=res_name, resource_type=res_type,
                                                    availability=res_ava, cost_use=res_cost_use, cost_unit=res_cost_unit,
                                                    resource_unit=res_unit)
        except XLSXParseError as exception:
            # rethrow same exception:
            raise
        except:
            # Any other non caught exception:
            raise XLSXParseError("""An error occurred while processing the Excel Resources sheet.\nCheck the input file if all necessary fields are entered.""")


        return resources_dict

    def process_risk_analysis(self, risk_analysis_sheet):
        """
        Reads an Excel Risk Analysis tab
        :return: dict: the risk analysis distributions in a dict, with as index the activity_id
        """
        risk_analysis_dict = {}
        col = 4
        for curr_row in range(self.get_nr_of_header_lines(risk_analysis_sheet), risk_analysis_sheet.get_highest_row()+1):
            # check if description is not None and not empty or only spaces
            if risk_analysis_sheet.cell(row=curr_row, column=col).value is not None and risk_analysis_sheet.cell(row=curr_row, column=col).value.strip():
                try:
                    risk_ana_dist_descript = risk_analysis_sheet.cell(row=curr_row, column=col).value.split("-")
                    if len(risk_ana_dist_descript) != 2:
                        raise XLSXParseError(("An error occurred while reading the Risk Analysis sheet on row {0} column {1}!\n" + \
                            "Expects a string in a format like  'manual - absolute' and not: {2}").format(curr_row, col, risk_analysis_sheet.cell(row=curr_row, column=col).value))

                    risk_ana_dist_type = risk_ana_dist_descript[0].strip()  # remove leading and trailing whitespace
                    risk_ana_dist_units = risk_ana_dist_descript[1].strip()  # remove leading and trailing whitespace
                    risk_ana_opt_duration = int(risk_analysis_sheet.cell(row=curr_row, column=col+1).value)
                    risk_ana_prob_duration = int(risk_analysis_sheet.cell(row=curr_row, column=col+2).value)
                    risk_ana_pess_duration = int(risk_analysis_sheet.cell(row=curr_row, column=col+3).value)
                    dict_id = int(risk_analysis_sheet.cell(row=curr_row, column=1).value)
                    risk_analysis_dict[dict_id] = RiskAnalysisDistribution(distribution_type=risk_ana_dist_type,
                                                                           distribution_units=risk_ana_dist_units,
                                                                           optimistic_duration=risk_ana_opt_duration,
                                                                           probable_duration=risk_ana_prob_duration,
                                                                           pessimistic_duration=risk_ana_pess_duration)
                except XLSXParseError:
                    # rethrow custom XLSXParseError exceptions
                    raise
                except:
                    raise XLSXParseError(("An error occurred while reading the Risk Analysis sheet on row {0}.\n" + \
                            "Check the input file and change row {0} to accomodate for the expected input format.").format(curr_row))

        return risk_analysis_dict

    def process_baseline_schedule(self, activities_sheet, resources_dict, risk_analysis_dict, agenda):
        """
        Reads an Excel Baseline Schedule tab
        :return: tuple: (activities_dict, activityGroups_dict, activityGroup_to_childActivities_dict, project_name)
        activities_dict: dict of all activities and activitygroups with 'key: value' = 'activityId: (row, Activity)'
        activityGroups_dict: dict of only activitygroups with 'key: value' = 'activityId: Activity'
        activityGroup_to_childActivities_dict: dict which links an activitygroupId to its child activities' ID, 'key: value' = 'activityId: [ActivityId1, ActivityId2,...]'
        project_name: string
        """
        activities_dict = {}  # contains all activities and group, activityId: (row, Activity) # TODO: why also store the row?
        activityGroups_dict = {} # contains only groups, activityId: Activity
        # process the first project line manually:
        first_data_row = self.get_nr_of_header_lines(activities_sheet)
        project_name = activities_sheet.cell(row=first_data_row, column=2).value
        # try to read the project WBS:
        project_wbs_str = activities_sheet.cell(row=first_data_row, column=3).value
        if project_wbs_str and str(project_wbs_str).strip():
            current_wbs = ()
            for number in str(project_wbs_str).split("."):
                if number: # check if non-empty string follows the '.'
                    current_wbs = current_wbs + (int(number),)
        else:
            # default project wbs when no wbs id given for the project:
            current_wbs = (1,)

        projectGroupActivity = Activity(0, name=project_name, wbs_id=current_wbs)
        activities_dict[0] = (first_data_row, projectGroupActivity)
        activityGroups_dict[0] = projectGroupActivity
        current_wbs = list(current_wbs) # current_wbs needs to be a list for further easy incrementing
        activityGroupFound = False # boolean for constructing auto generated WBS

        # loop over all rows, except the project row parsed before
        for curr_row in range(first_data_row + 1, activities_sheet.get_highest_row()+1):
            activity_id = -1
            try:
                # try to cast this string to an int
                activity_id = int(activities_sheet.cell(row=curr_row, column=1).value)
            except:
                # This row has no valid ID number => skip this row
                print("XLSXparser:process_baseline_schedule: Invalid activity ID at row {0}".format(curr_row))
                continue
            
            activity_name = activities_sheet.cell(row=curr_row, column=2).value
            activity_wbs = ()
            # check if WBS field is not None and not empty or only spaces:
            if activities_sheet.cell(row=curr_row, column=3).value and str(activities_sheet.cell(row=curr_row, column=3).value).strip():
                # WBS id given:
                for number in str(activities_sheet.cell(row=curr_row, column=3).value).split("."):
                    if number:
                        activity_wbs = activity_wbs + (int(number),)
                # update current_wbs:
                current_wbs = list(activity_wbs)
            else:
                # auto generate WBS id:
                # check if is activityGroup by var_cost: # NOTE: this is hack to determine if this row belongs to an activityGroup
                if activities_sheet.cell(row=curr_row, column=13).value is None:
                    # current row belongs to an activityGroup
                    # don't read any further, calculate its aggregate value later on:
                    if len(current_wbs) == 2:
                        current_wbs[1] += 1
                    elif len(current_wbs) > 2:
                        current_wbs = current_wbs[:2]
                        current_wbs[1] += 1
                    else:
                        # previous node was the project root:
                        current_wbs.append(1)

                    activityGroupFound = True
                    activityGroup = Activity(activity_id, name=activity_name, wbs_id=tuple(current_wbs))
                    activityGroups_dict[activity_id] = activityGroup
                    activities_dict[activity_id] = (curr_row, activityGroup)
                    continue
                else:
                    # construct wbs of an activity:
                    if len(current_wbs) == 2:
                        if not activityGroupFound:
                            current_wbs[1] += 1
                        else:
                            current_wbs.append(1)
                    elif len(current_wbs) > 2:
                        current_wbs[-1] += 1
                    else:
                        # previous node was the project root:
                        current_wbs.append(1)
                activity_wbs = tuple(current_wbs)

            # check if is activityGroup by checking if it has a var_cost
            if activities_sheet.cell(row=curr_row, column=13).value is None:
                # current row belongs to activityGroup; don't read any further, calculate its aggregate value later on
                activityGroupFound = True
                activityGroup = Activity(activity_id, name=activity_name, wbs_id=activity_wbs)
                activityGroups_dict[activity_id] = activityGroup
                activities_dict[activity_id] = (curr_row, activityGroup)
                continue

            activity_predecessors = None
            activity_successors = None
            try:
                activity_predecessors = self.process_predecessors(activities_sheet.cell(row=curr_row, column=4).value, agenda)
            except:
                raise XLSXParseError(("An error occurred while reading the predecessors of row {0} with value: {1}\n" + \
                            "Check if the input format is valid.").format(curr_row, activities_sheet.cell(row=curr_row, column=4).value))
            try:
                activity_successors = self.process_successors(activities_sheet.cell(row=curr_row, column=5).value, agenda)
            except:
                raise XLSXParseError(("An error occurred while reading the successors of row {0} with value: {1}\n" + \
                            "Check if the input format is valid.").format(curr_row, activities_sheet.cell(row=curr_row, column=5).value))

            if activities_sheet.cell(row=curr_row, column=6).value:
                baseline_start = self.read_date(activities_sheet.cell(row=curr_row, column=6).value)
            else:
                raise XLSXParseError("No baseline start date was given on row {0} of the baseline schedule.".format(curr_row))

            # check if duration is given to calculate baseline end:
            baseline_duration = 0
            baseline_end = datetime.datetime.max
            # read the field values of the mutual alternative fields:
            baseline_end_field = activities_sheet.cell(row=curr_row, column=7).value
            baseline_duration_field = activities_sheet.cell(row=curr_row, column=8).value

            if baseline_duration_field and baseline_duration_field.strip(): # if not empty or not only spaces
                # baseline duration is given
                baseline_duration_workingHours = agenda.convert_durationString_to_workingHours(baseline_duration_field)
                baseline_duration = agenda.get_workingDuration_timedelta(baseline_duration_workingHours)
                # calculate baseline end:
                baseline_end = agenda.get_end_date(baseline_start, baseline_duration.days, baseline_duration_workingHours - baseline_duration.days * agenda.get_working_hours_in_a_day())
            elif baseline_end_field and str(baseline_end_field).strip(): # if not empty or not only spaces
                baseline_end = self.read_date(baseline_end_field)
                # calculate baseline duration:
                baseline_duration = agenda.get_time_between(baseline_start, baseline_end)
                baseline_duration_workingHours = agenda.get_workingDuration_workingHours(baseline_duration)
            else:
                raise XLSXParseError("process_baseline_schedule:Activity on row {0}, has no basline duration or baseline end!".format(curr_row))

            # process resource demand:
            activity_resources = []
            activity_resource_cost = 0.0 # calculate resource cost instead of reading it
            if activities_sheet.cell(row=curr_row, column=9).value:
                activity_resource_cost = self.process_activity_resource_assignments(activities_sheet.cell(row=curr_row, column=9).value, activity_resources, resources_dict, baseline_duration_workingHours)

            baseline_fixed_cost = 0.0
            if activities_sheet.cell(row=curr_row, column=11).value and str(activities_sheet.cell(row=curr_row, column=11).value).strip():
                baseline_fixed_cost = float(activities_sheet.cell(row=curr_row, column=11).value)

            # check if cost/hour or variable cost is given to calculate values:
            baseline_hourly_cost = 0.0
            baseline_var_cost = 0.0

            baseline_hourly_cost_field = activities_sheet.cell(row=curr_row, column=12).value
            baseline_var_cost_field = activities_sheet.cell(row=curr_row, column=13).value

            # first check if var_cost is given, else try to use the hourly cost:
            if baseline_var_cost_field is not None and str(baseline_var_cost_field).strip():
                baseline_var_cost = float(baseline_var_cost_field)
                # calculate hourly cost:
                baseline_hourly_cost = baseline_var_cost / baseline_duration_workingHours if baseline_duration_workingHours != 0 else 0

            elif baseline_hourly_cost_field is not None and str(baseline_hourly_cost_field).strip():
                baseline_hourly_cost = float(baseline_hourly_cost_field)
                # calculate var_cost:
                baseline_var_cost = baseline_hourly_cost * baseline_duration_workingHours
            else:
                raise XLSXParseError("process_baseline_schedule:Acitivity on row {0}, has no baseline variable cost or baseline hourly cost given!".format(curr_row))

            # calculate total_cost:
            baseline_total_cost = baseline_fixed_cost + baseline_var_cost + activity_resource_cost

            activity_baseline_schedule = BaselineScheduleRecord(start=baseline_start, end=baseline_end,
                                                                duration=baseline_duration,
                                                                fixed_cost=baseline_fixed_cost,
                                                                hourly_cost=baseline_hourly_cost,
                                                                var_cost=baseline_var_cost,
                                                                total_cost=baseline_total_cost)
            activity_risk_analysis = None
            if activity_id in risk_analysis_dict:
                activity_risk_analysis = risk_analysis_dict[activity_id]


            activities_dict[activity_id] = (curr_row, Activity(activity_id, name=activity_name, wbs_id=activity_wbs,
                                               predecessors=activity_predecessors, successors=activity_successors,
                                               baseline_schedule=activity_baseline_schedule,
                                               resource_cost=activity_resource_cost,
                                               risk_analysis=activity_risk_analysis, resources=activity_resources))
        # Recheck if read activityGroups have subactivities and if activities don't have subactivities:
        Activity.check_lists_activities_groups(activities_dict, activityGroups_dict) # adds activities which appear to have subactivities to the activityGroups

        # get a dict, linking each activityId to a list of all its child activityId's:
        activityGroups = list(activityGroups_dict.values())
        activitiesOnly_dict = {activity_id: x[1] for activity_id, x in activities_dict.items() if activity_id not in activityGroups_dict}
        activityGroup_to_childActivities_dict = Activity.generate_activityGroups_to_childActivities_dict(list(activitiesOnly_dict.values()), activityGroups)
        Activity.update_activityGroups_aggregated_values(activityGroups, activitiesOnly_dict, activityGroup_to_childActivities_dict, agenda)

        return activities_dict, activityGroups_dict, activityGroup_to_childActivities_dict, project_name

    """ WRITING PRIVATE FUNCTIONS """
    def get_all_activities(self, tracking_period):
        activities = []
        for atr in tracking_period.tracking_period_records:
            activities.append(atr.activity)
        return activities

    def calculate_aggregated_keyMetrics(self, tracking_period):
        sum_pv = 0.0
        sum_ev = 0.0
        sum_ac = 0.0
        for atr in tracking_period.tracking_period_records:
            if not Activity.is_not_lowest_level_activity(atr.activity, self.get_all_activities(tracking_period)):
                sum_pv += atr.planned_value
                sum_ev += atr.earned_value
                sum_ac += atr.actual_cost
        return sum_pv, sum_ev, sum_ac

    def calculate_eac(self, ac, bac, ev, pf):
        if pf == 0:
            return 0
        return ac + (bac - ev)/float(pf)

    def get_pv(self, tracking_period):
        for atr in tracking_period.tracking_period_records:
            if len(atr.activity.wbs_id) == 1:
                return atr.planned_value
        return None

    def calculate_es(self, project_object, PVcurve, current_EV, current_PV, currentTime):
        """
        This function calculates the ES datetime based on given PVcurve, current EV value and current time

        :param project_object: ProjectObject, the project object corresponding to the PVcurve
        :param PVcurve: list of tuples, (PV cumsum value, datetime of this PV cumsum value) as calculated by calculate_PVcurve
        :param current_EV: float, Earned value for which to search the ES
        :param current_PV: float, Planned Value at the current statusdate
        :param currentTime: datetime, statusdate
        """
        # algorithm:
        # if EV = PV(statusdate): ES = statusdate
        # elif EV < PV(statusdate):
            # Find t such that EV >= PV(t) and EV < PV(t+1)
            # ES = t
        # else EV > PV(statusdate):
            # Find t such that EV > PV(t-1) and EV <= PV(t)
            # ES = t

        ## according to PMKnowledgecenter the following line should be used, but an ES in the middle of non-working datetimes is not desirable!
        ## ES = t + (EV - PV(t)) / (PV(t+1) - PV(t)) * (next_t - t)

        # if EV = PV(statusdate): ES = statusdate
        if abs(current_EV - current_PV) < 1e-3:
            # check if currentTime is not later than latest baseline date:
            projectBaslineEndDate = max([activity.baseline_schedule.end for activity in project_object.activities]) # projectBaselineEndDate
            if currentTime <= projectBaslineEndDate:
                # make sure currentTime is a valid endTime:
                return project_object.agenda.get_end_date(currentTime, 0,0)
            else:
                # ES can't be larger than the scheduled project end date:
                return projectBaslineEndDate

        t = min([activity.baseline_schedule.start for activity in project_object.activities])  # projectBaselineStartDate
        lowerPV = -1
        pointFound = False

        searchFirst_EVdate = current_EV > current_PV # boolean indicating which point on the PV curve to search: the first point where EV = PV or the last point

        # search first PV which is larger than EV
        for i in range(1, len(PVcurve)):
            if searchFirst_EVdate and (abs(PVcurve[i][0] - current_EV) < 1e-3 or PVcurve[i][0] >= current_EV):
                if abs(PVcurve[i][0] - current_EV) < 1e-3:
                    # first PV point found which is equal to the given EV value
                    t = PVcurve[i][1]
                else:
                    # found PV(i) is already larger than EV => take the previous PV(i-1)
                    t = PVcurve[i-1][1]
                pointFound = True
                break
            elif not searchFirst_EVdate and PVcurve[i][0] > current_EV and abs(PVcurve[i][0] - current_EV) >= 1e-3:
                # first PV point found which is larger than the given EV value
                t = PVcurve[i-1][1]
                lowerPV = PVcurve[i-1][0]
                pointFound = True
                break
        #endFor searching larger PV point

        if not pointFound:
            t = max([activity.baseline_schedule.end for activity in project_object.activities])  # projectBaselineEndDate

        return t

    def calculate_PVcurve(self, project_object):
        """
        This function generates the PV curve of the baseline schedule of the given project.
        return: generated_PVcurve: list of tuples: [(PV, datetime), ...]
        """

        # retrieve only lowest level activities and sort them on their baseline start date and end date
        lowestLevelActivities = [activity for activity in project_object.activities if not Activity.is_not_lowest_level_activity(activity, project_object.activities)]
        lowestLevelActivitiesSorted = sorted(lowestLevelActivities, key=attrgetter('baseline_schedule.start', 'baseline_schedule.end'))

        projectBaselineStartDate = min([activity.baseline_schedule.start for activity in project_object.activities])
        projectBaselineEndDate = max([activity.baseline_schedule.end for activity in project_object.activities])

        # initial starting point
        generated_PVcurve = [(0, projectBaselineStartDate)]

        # search end of first working hour of this project
        currentDatetime = project_object.agenda.get_next_date(projectBaselineStartDate, 0, 0) + datetime.timedelta(hours = 1)

        # loop variables
        index_NextActivityToAdd = 0
        currentPVcumsumValue = 0
        runningActivities = []

        # traverse the complete project baseline duration
        while currentDatetime <= projectBaselineEndDate:
            # update list of running activities at this time:
            # remove ended activities: activities that finished before the start of the evaluating interval hour that ended on currentDatetime
            for activity in runningActivities.copy():
                if activity.baseline_schedule.end < currentDatetime:
                    # activity ended maximally on the start of the currently evaluating interval hour of currentDatetime
                    runningActivities.remove(activity)

            # add activities that are just started:
            while (index_NextActivityToAdd < len(lowestLevelActivitiesSorted)) and lowestLevelActivitiesSorted[index_NextActivityToAdd].baseline_schedule.start < currentDatetime:
                runningActivities.append(lowestLevelActivitiesSorted[index_NextActivityToAdd])

                # add starting fixed cost of the newly added activity to the PVcumsumValue:
                currentPVcumsumValue += lowestLevelActivitiesSorted[index_NextActivityToAdd].baseline_schedule.fixed_cost
                # add starting use cost of used resources by this activity:
                for resourceTuple in lowestLevelActivitiesSorted[index_NextActivityToAdd].resources:
                    currentPVcumsumValue +=  Resource.calculate_resource_assignment_cost(resourceTuple[0], resourceTuple[1], resourceTuple[2], 0)
                #endFor adding resource use costs
                index_NextActivityToAdd += 1

            # update PVcurve with one timestemp its variable added value:
            for activity in runningActivities:
                currentPVcumsumValue += activity.baseline_schedule.hourly_cost
                # add also variable cost of the use of its resources
                for resourceTuple in activity.resources:
                    if not resourceTuple[2]:
                        # if the resource assignment is not fixed:
                        currentPVcumsumValue += resourceTuple[0].cost_unit * resourceTuple[1]  # cost/(hour * unit) * demanded units

            # save the currentPVcumsumValue at its new timestep:
            generated_PVcurve.append((currentPVcumsumValue, currentDatetime))

            # advance to next working hour its end
            currentDatetime = project_object.agenda.get_next_date(currentDatetime, 0, 0) + datetime.timedelta(hours = 1)

        return generated_PVcurve

    def calculate_SVt(self, project_object, ES, currentTime):
        """This function calculates the SV(t) in hours, based on the given ES and currentTime.
        returns: (SV(t) in workinghours, string representation of SV(t))
        """
        # determine the time between ES and currentTime:

        if ES >= currentTime:
            timeBetween = project_object.agenda.get_time_between(currentTime, ES)
            return (timeBetween.days * project_object.agenda.get_working_hours_in_a_day() + round(timeBetween.seconds / 3600), XLSXParser.get_duration_str(timeBetween, False))
        else:
            timeBetween = project_object.agenda.get_time_between(ES, currentTime)
            return (- timeBetween.days * project_object.agenda.get_working_hours_in_a_day() - round(timeBetween.seconds / 3600), XLSXParser.get_duration_str(timeBetween, True))

    def calculate_SPIt(self, project_object, ES, currentTime):
        "This fuction calculates the SPI(t) value = ES / AT"

        projectBaselineStartDate = min([activity.baseline_schedule.start for activity in project_object.activities])
        durationWorkingTimeES = project_object.agenda.get_time_between(projectBaselineStartDate , ES)
        durationWorkingTimeAT = project_object.agenda.get_time_between(projectBaselineStartDate , currentTime)

        if durationWorkingTimeAT.total_seconds() <= 0:
            return 0.0

        return (durationWorkingTimeES.days * project_object.agenda.get_working_hours_in_a_day() + round(durationWorkingTimeES.seconds / 3600)) / (durationWorkingTimeAT.days * project_object.agenda.get_working_hours_in_a_day() + round(durationWorkingTimeAT.seconds / 3600.0))

    def calculate_p_factor(self, project_object, tracking_period, ES):
        "This function calculates the p-factor for the given tracking_period and ES datetime"
        # using algorithm:
        # p-factor = sum(i:N)[ min(PV(i,ES), EV(i,AT))] / sum(i:N)[ PV(i,ES)]
        # with
        # N: The set of all activities of the project network
        # PV(i,ES): The planned value of activity i at time ES
        # EV(i,AT): The earned value of activity i at time AT

        # sort the tracking_period_records according to its activity baseline schedule start and end date => can figure out which activity uses first a consumable resource

        # retrieve only lowest level activities and sort them on their baseline start date and end date
        lowestLevelActivitiesTrackingRecords = [atr for atr in tracking_period.tracking_period_records if not Activity.is_not_lowest_level_activity(atr.activity, self.get_all_activities(tracking_period))]
        lowestLevelActivitiesTrackingRecordsSorted = sorted(lowestLevelActivitiesTrackingRecords, key=attrgetter('activity.baseline_schedule.start', 'activity.baseline_schedule.end'))

        numerator = 0.0
        denominator = 0.0

        for atr in lowestLevelActivitiesTrackingRecordsSorted:
            # determine The planned value of activity i at time ES:
            PV_i_ES = 0
            if ES <= atr.activity.baseline_schedule.start:
                # activity is not yet started according to plan => no planned value
                pass
            elif ES >= atr.activity.baseline_schedule.end:
                # activity is already finished according to plan => PVi = total_cost
                PV_i_ES = atr.activity.baseline_schedule.total_cost
            else:
                # activity is still running => calculate intermediate PV value:
                # add variable costs of this activity according to duration that this activity is running:
                # determine running duration:
                runningTimedelta = project_object.agenda.get_time_between(atr.activity.baseline_schedule.start, ES)
                runningWorkingHours = runningTimedelta.days * project_object.agenda.get_working_hours_in_a_day() + int(runningTimedelta.seconds / 3600)

                PV_i_ES = atr.activity.baseline_schedule.fixed_cost + atr.activity.baseline_schedule.hourly_cost * runningWorkingHours

                # add starting use cost of used resources by this activity:
                for resourceTuple in atr.activity.resources:
                    PV_i_ES += Resource.calculate_resource_assignment_cost(resourceTuple[0], resourceTuple[1], resourceTuple[2], runningWorkingHours)
                #endFor adding resource costs

            #endIF calculating PV_i_ES
            numerator += min(PV_i_ES, atr.earned_value)
            denominator += PV_i_ES
        #endFor all activity tracking records

        if denominator < 1e-10:
            return 0.0
        else:
            return numerator / denominator

    """ ---------------------- MAIN FUNCTION TO WRITE TO EXCEL ------------------------- """

    def from_schedule_object(self, project_object, file_path_output):
        """
        This is just a lot of writing to excel code, it is ugly..

        """
        workbook = xlsxwriter.Workbook(file_path_output)

        # Lots of formats
        header = workbook.add_format({'bold': True, 'bg_color': '#316AC5', 'font_color': 'white', 'text_wrap': True,
                                      'border': 1, 'font_size': 8})
        yellow_cell = workbook.add_format({'bg_color': '#FFFF00', 'text_wrap': True, 'border': 1, 'font_size': 8})
        green_cell = workbook.add_format({'bg_color': '#9BBB59', 'text_wrap': True, 'border': 1, 'font_size': 8})
        red_cell_agenda = workbook.add_format({'bg_color': '#FFC7CE', 'text_wrap': True, 'border': 1, 'font_size': 8, 'font_color': '#9C0006'})
        green_cell_agenda = workbook.add_format({'bg_color': '#C6EFCE', 'text_wrap': True, 'border': 1, 'font_size': 8, 'font_color': '#006100'})
        gray_cell = workbook.add_format({'bg_color': '#D4D0C8', 'text_wrap': True, 'border': 1, 'font_size': 8})
        blue_cell = workbook.add_format({'bg_color': '#D9EAF7', 'text_wrap': True, 'border': 1, 'font_size': 8})
        
        date_green_cell = workbook.add_format({'bg_color': '#9BBB59', 'text_wrap': True, 'border': 1, 'num_format': 'dd/mm/yyyy H:MM', 'font_size': 8})
        date_gray_cell = workbook.add_format({'bg_color': '#D4D0C8', 'text_wrap': True, 'border': 1, 'num_format': 'dd/mm/yyyy H:MM', 'font_size': 8})
        date_blue_cell = workbook.add_format({'bg_color': '#D9EAF7', 'text_wrap': True, 'border': 1, 'num_format': 'dd/mm/yyyy H:MM', 'font_size': 8})
        holiday_yellow_cell = workbook.add_format({'bg_color': '#FFFF00', 'text_wrap': True, 'border': 1, 'num_format': 'dd/mm/yyyy', 'font_size': 8,})
        
        money_green_cell = workbook.add_format({'bg_color': '#9BBB59', 'text_wrap': True, 'border': 1, 'num_format': '#,##0.00' + u"\u20AC", 'font_size': 8})
        money_yellow_cell = workbook.add_format({'bg_color': '#FFFF00', 'text_wrap': True, 'border': 1, 'num_format': '#,##0.00' + u"\u20AC", 'font_size': 8})
        money_gray_cell = workbook.add_format({'bg_color': '#D4D0C8', 'text_wrap': True, 'border': 1, 'num_format': '#,##0.00' + u"\u20AC", 'font_size': 8})
        money_blue_cell = workbook.add_format({'bg_color': '#D9EAF7', 'text_wrap': True, 'border': 1, 'num_format': '#,##0.00' + u"\u20AC", 'font_size': 8})

        percent_green_cell = workbook.add_format({'bg_color': '#9BBB59', 'text_wrap': True, 'border': 1, 'font_size': 8,'num_format': '0%'})
        percent_gray_cell = workbook.add_format({'bg_color': '#D4D0C8', 'text_wrap': True, 'border': 1, 'font_size': 8,'num_format': '0%'})
        percent_blue_cell = workbook.add_format({'bg_color': '#D9EAF7', 'text_wrap': True, 'border': 1, 'font_size': 8,'num_format': '0%'})
        

        # Worksheets
        bsch_worksheet = workbook.add_worksheet("Baseline Schedule")
        res_worksheet = workbook.add_worksheet("Resources")
        ra_worksheet = workbook.add_worksheet("Risk Analysis")

        # Write the Baseline Schedule Worksheet

        # Set the width of the columns
        bsch_worksheet.set_column(0, 0, 3)
        bsch_worksheet.set_column(1, 1, 25)
        bsch_worksheet.set_column(2, 2, 5)
        bsch_worksheet.set_column(3, 4, 16)
        bsch_worksheet.set_column(5, 6, 13)
        bsch_worksheet.set_column(7, 7, 8)
        bsch_worksheet.set_column(8, 8, 25)
        bsch_worksheet.set_column(9, 9, 10)
        bsch_worksheet.set_column(10, 11, 10)
        bsch_worksheet.set_column(13, 13, 12)

        # Set the height of rows
        bsch_worksheet.set_row(1, 30)

        # Write header cells (using the header format, and by merging some cells)
        bsch_worksheet.merge_range('A1:C1', "General", header)
        bsch_worksheet.merge_range('D1:E1', "Relations", header)
        bsch_worksheet.merge_range('F1:H1', "Baseline", header)
        bsch_worksheet.merge_range('I1:J1', "Resource Demand", header)
        bsch_worksheet.merge_range('K1:N1', "Baseline Costs", header)

        bsch_worksheet.write('A2', "ID", header)
        bsch_worksheet.write('B2', "Name", header)
        bsch_worksheet.write('C2', "WBS", header)
        bsch_worksheet.write('D2', "Predecessors", header)
        bsch_worksheet.write('E2', "Successors", header)
        bsch_worksheet.write('F2', "Baseline Start", header)
        bsch_worksheet.write('G2', "Baseline End", header)
        bsch_worksheet.write('H2', "Duration", header)
        bsch_worksheet.write('I2', "Resource Demand", header)
        bsch_worksheet.write('J2', "Resource Cost", header)
        bsch_worksheet.write('K2', "Fixed Cost", header)
        bsch_worksheet.write('L2', "Cost/Hour", header)
        bsch_worksheet.write('M2', "Variable Cost", header)
        bsch_worksheet.write('N2', "Total Cost", header)

        # Now we run through all activities to get the required information
        counter = 2
        for activity in project_object.activities:
            if not Activity.is_not_lowest_level_activity(activity, project_object.activities):
                # Write activity of lowest level
                bsch_worksheet.write_number(counter, 0, activity.activity_id, green_cell)
                bsch_worksheet.write(counter, 1, str(activity.name), green_cell)
                self.write_wbs(bsch_worksheet, counter, 2, activity.wbs_id, yellow_cell)
                self.write_predecessors(bsch_worksheet, counter, 3, activity.predecessors, green_cell, project_object.agenda)
                self.write_successors(bsch_worksheet, counter, 4, activity.successors, green_cell, project_object.agenda)
                bsch_worksheet.write_datetime(counter, 5, activity.baseline_schedule.start, date_green_cell)
                bsch_worksheet.write_datetime(counter, 6, activity.baseline_schedule.end, date_gray_cell)
                bsch_worksheet.write(counter, 7, self.get_duration_str(activity.baseline_schedule.duration), green_cell)
                self.write_resources(bsch_worksheet, counter, 8, activity.resources, yellow_cell)
                bsch_worksheet.write_number(counter, 9, activity.resource_cost, money_gray_cell)
                bsch_worksheet.write_number(counter, 10, activity.baseline_schedule.fixed_cost, money_green_cell)
                bsch_worksheet.write_number(counter, 11, activity.baseline_schedule.hourly_cost, money_gray_cell)
                if activity.baseline_schedule.var_cost is not None:
                    bsch_worksheet.write_number(counter, 12, activity.baseline_schedule.var_cost, money_green_cell)
                else:
                    bsch_worksheet.write_number(counter, 12, 0, money_green_cell)
                bsch_worksheet.write_number(counter, 13, activity.baseline_schedule.total_cost, money_gray_cell)
            else:
                # Write aggregated activity
                if activity.activity_id == 0:
                    # project group:
                    bsch_worksheet.write_number(counter, 0, activity.activity_id, blue_cell)
                    bsch_worksheet.write(counter, 1, str(activity.name), green_cell)
                else:
                    bsch_worksheet.write_number(counter, 0, activity.activity_id, yellow_cell)
                    bsch_worksheet.write(counter, 1, str(activity.name), yellow_cell)
                self.write_wbs(bsch_worksheet, counter, 2, activity.wbs_id, yellow_cell)
                self.write_predecessors(bsch_worksheet, counter, 3, activity.predecessors, blue_cell, project_object.agenda) # normally empty
                self.write_successors(bsch_worksheet, counter, 4, activity.successors, blue_cell, project_object.agenda) # normally empty
                bsch_worksheet.write_datetime(counter, 5, activity.baseline_schedule.start, date_blue_cell)
                bsch_worksheet.write_datetime(counter, 6, activity.baseline_schedule.end, date_blue_cell)
                bsch_worksheet.write(counter, 7, self.get_duration_str(activity.baseline_schedule.duration), blue_cell)
                self.write_resources(bsch_worksheet, counter, 8, activity.resources, blue_cell) # normally empty
                bsch_worksheet.write(counter, 9, "", money_blue_cell)
                bsch_worksheet.write_number(counter, 10, activity.baseline_schedule.fixed_cost, money_blue_cell)
                bsch_worksheet.write(counter, 11, "", money_blue_cell)
                bsch_worksheet.write(counter, 12, "", money_blue_cell)
                bsch_worksheet.write_number(counter, 13, activity.baseline_schedule.total_cost, money_blue_cell)

            counter += 1

        # Write the resources sheet

        # Some small adjustments to rows and columns in the resource sheet
        res_worksheet.set_row(1, 25)
        res_worksheet.set_column(1, 1, 15)
        res_worksheet.set_column(6, 6, 40)
        res_worksheet.set_column(7, 7, 10)

        # Write header cells (using the header format, and by merging some cells)
        res_worksheet.merge_range('A1:D1', "General", header)
        res_worksheet.merge_range('E1:F1', "Resource Cost", header)
        res_worksheet.merge_range('G1:H1', "Resource Demand", header)

        res_worksheet.write('A2', "ID", header)
        res_worksheet.write('B2', "Name", header)
        res_worksheet.write('C2', "Type", header)
        res_worksheet.write('D2', "Availability", header)
        res_worksheet.write('E2', "Cost/Use", header)
        res_worksheet.write('F2', "Cost/Unit", header)
        res_worksheet.write('G2', "Assigned To", header)
        res_worksheet.write('H2', "Total Cost", header)

        counter = 2
        for resource in project_object.resources:
            res_worksheet.write_number(counter, 0, resource.resource_id, yellow_cell)
            res_worksheet.write(counter, 1, resource.name, yellow_cell)
            res_worksheet.write(counter, 2, resource.resource_type.value, yellow_cell)
            # God knows why we write the availability twice, it was like that in the template
            useless_availability_string = str(resource.availability) + " " + str(resource.resource_unit)
            res_worksheet.write(counter, 3, useless_availability_string, yellow_cell)
            res_worksheet.write(counter, 4, resource.cost_use, money_yellow_cell)
            res_worksheet.write(counter, 5, resource.cost_unit, money_yellow_cell)
            self.write_resource_assign_cost(res_worksheet, counter, 6, resource, project_object.activities, gray_cell,
                                            money_gray_cell, project_object)
            counter += 1

        # Write the risk analysis sheet

        # Adjust some column widths
        ra_worksheet.set_column(0, 0, 5)
        ra_worksheet.set_column(1, 1, 18)
        ra_worksheet.set_column(3, 3, 15)
        ra_worksheet.set_column(4, 6, 12)

        # Write the headers
        ra_worksheet.merge_range('A1:B1', "General", header)
        ra_worksheet.write('C1', "Baseline", header)
        ra_worksheet.merge_range('D1:G1', "Activity Duration Distribution Profiles", header)

        ra_worksheet.write('A2', "ID", header)
        ra_worksheet.write('B2', "Name", header)
        ra_worksheet.write('C2', "Duration", header)
        ra_worksheet.write('D2', "Description", header)
        ra_worksheet.write('E2', "Optimistic", header)
        ra_worksheet.write('F2', "Most Probable", header)
        ra_worksheet.write('G2', "Pessimistic", header)

        # Write the rows by iterating through the activities (since they are linked to it)
        counter = 2
        for activity in project_object.activities:
            if Activity.is_not_lowest_level_activity(activity, project_object.activities):
                # activity group
                if activity.activity_id == 0:
                    # Project Group:
                    ra_worksheet.write_number(counter, 0, activity.activity_id, blue_cell)
                    ra_worksheet.write(counter, 1, str(activity.name), green_cell)

                else:
                    # std activity group:
                    ra_worksheet.write_number(counter, 0, activity.activity_id, yellow_cell)
                    ra_worksheet.write(counter, 1, str(activity.name), yellow_cell)

                ra_worksheet.write(counter, 2, self.get_duration_hours_str(activity.baseline_schedule.duration, project_object.agenda.get_working_hours_in_a_day()), blue_cell)
                ra_worksheet.write(counter, 3, "", blue_cell)
                ra_worksheet.write(counter, 4, "", blue_cell)
                ra_worksheet.write(counter, 5, "", blue_cell)
                ra_worksheet.write(counter, 6, "", blue_cell)
            else:
                # activity
                ra_worksheet.write_number(counter, 0, activity.activity_id, green_cell)
                ra_worksheet.write(counter, 1, str(activity.name), green_cell)
                ra_worksheet.write(counter, 2, self.get_duration_hours_str(activity.baseline_schedule.duration, project_object.agenda.get_working_hours_in_a_day()), gray_cell)
                description = str(activity.risk_analysis.distribution_type.value) + " - " \
                                  + str(activity.risk_analysis.distribution_units.value)
                ra_worksheet.write(counter, 3, description, yellow_cell)
                ra_worksheet.write_number(counter, 4, activity.risk_analysis.optimistic_duration, yellow_cell)
                ra_worksheet.write_number(counter, 5, activity.risk_analysis.probable_duration, yellow_cell)
                ra_worksheet.write_number(counter, 6, activity.risk_analysis.pessimistic_duration, yellow_cell)
            counter += 1

        # Write the tracking periods, same drill, multiple sheets, too many bloody columns
        for i in range(0, len(project_object.tracking_periods)):
            if i == 0:
                tracking_period_worksheet = workbook.add_worksheet("Project Control - TP1")
            else:
                tracking_period_worksheet = workbook.add_worksheet("TP" + str(i+1))

            # Set column widths and create headers
            tracking_period_worksheet.set_column(0, 0, 5)
            tracking_period_worksheet.set_column(1, 1, 18)
            tracking_period_worksheet.set_column(2, 3, 13)
            tracking_period_worksheet.set_column(4, 4, 6)
            tracking_period_worksheet.set_column(5, 5, 22)
            tracking_period_worksheet.set_column(6, 10, 10)
            tracking_period_worksheet.set_column(11, 11, 13)
            tracking_period_worksheet.set_column(12, 12, 6)
            tracking_period_worksheet.set_column(13, 20, 10)
            tracking_period_worksheet.set_column(21, 21, 8)
            tracking_period_worksheet.set_column(22, 23, 10)
            tracking_period_worksheet.set_row(3, 30)

            tracking_period_worksheet.write('B1', 'TP Status Date', header)
            tracking_period_worksheet.write('E1', 'TP Name', header)
            tracking_period_worksheet.merge_range('A3:B3', "General", header)
            tracking_period_worksheet.merge_range('C3:E3', "Baseline", header)
            tracking_period_worksheet.merge_range('F3:G3', "Resource Demand", header)
            tracking_period_worksheet.merge_range('H3:K3', "Baseline Costs", header)
            tracking_period_worksheet.merge_range('L3:X3', "Tracking", header)

            tracking_period_worksheet.write('A4', 'ID', header)
            tracking_period_worksheet.write('B4', 'Name', header)
            tracking_period_worksheet.write('C4', 'Baseline Start', header)
            tracking_period_worksheet.write('D4', 'Baseline End', header)
            tracking_period_worksheet.write('E4', 'Duration', header)
            tracking_period_worksheet.write('F4', 'Resource Demand', header)
            tracking_period_worksheet.write('G4', 'Resource Cost', header)
            tracking_period_worksheet.write('H4', 'Fixed Cost', header)
            tracking_period_worksheet.write('I4', 'Cost/Hour', header)
            tracking_period_worksheet.write('J4', 'Variable Cost', header)
            tracking_period_worksheet.write('K4', 'Total Cost', header)
            tracking_period_worksheet.write('L4', 'Actual Start', header)
            tracking_period_worksheet.write('M4', 'Actual Duration', header)
            tracking_period_worksheet.write('N4', 'PAC', header)
            tracking_period_worksheet.write('O4', 'PRC', header)
            tracking_period_worksheet.write('P4', 'Remaining Duration', header)
            tracking_period_worksheet.write('Q4', 'PAC Dev', header)
            tracking_period_worksheet.write('R4', 'PRC Dev', header)
            tracking_period_worksheet.write('S4', 'Actual Cost', header)
            tracking_period_worksheet.write('T4', 'Remaining Cost', header)
            tracking_period_worksheet.write('U4', 'Percentage Completed', header)
            tracking_period_worksheet.write('V4', 'Tracking', header)
            tracking_period_worksheet.write('W4', 'Earned Value (EV)', header)
            tracking_period_worksheet.write('X4', 'Planned Value (PV)', header)

            # Write the data
            tracking_period_worksheet.write_datetime('C1', project_object.tracking_periods[i].tracking_period_statusdate
                                                     , date_green_cell)
            tracking_period_worksheet.write('F1', project_object.tracking_periods[i].tracking_period_name,
                                                     yellow_cell)
            counter = 4

            for atr in project_object.tracking_periods[i].tracking_period_records:  # atr = ActivityTrackingRecord
                if Activity.is_not_lowest_level_activity(atr.activity, project_object.activities):
                    # work package here:
                    if atr.activity.activity_id == 0:
                        # project group:
                        tracking_period_worksheet.write_number(counter, 0, atr.activity.activity_id, blue_cell)
                        tracking_period_worksheet.write(counter, 1, atr.activity.name, green_cell)
                    else:
                        tracking_period_worksheet.write_number(counter, 0, atr.activity.activity_id, yellow_cell)
                        tracking_period_worksheet.write(counter, 1, atr.activity.name, yellow_cell)

                    tracking_period_worksheet.write_datetime(counter, 2, atr.activity.baseline_schedule.start, date_blue_cell)
                    tracking_period_worksheet.write_datetime(counter, 3, atr.activity.baseline_schedule.end, date_blue_cell)
                    tracking_period_worksheet.write(counter, 4, self.get_duration_str(atr.activity.baseline_schedule.duration), blue_cell)
                    tracking_period_worksheet.write(counter, 5, "", blue_cell) # not printed for workpackages
                    tracking_period_worksheet.write(counter, 6, "", money_blue_cell) # not printed for workpackages
                    tracking_period_worksheet.write_number(counter, 7, atr.activity.baseline_schedule.fixed_cost, money_blue_cell)
                    tracking_period_worksheet.write(counter, 8, "", money_blue_cell) # not printed for workpackages
                    tracking_period_worksheet.write(counter, 9, "", money_blue_cell)
                    tracking_period_worksheet.write_number(counter, 10, atr.activity.baseline_schedule.total_cost, money_blue_cell)

                    tracking_period_worksheet.write(counter, 11, '', blue_cell) # not printed for workpackages
                    tracking_period_worksheet.write(counter, 12, self.get_duration_str(atr.actual_duration), blue_cell)
                    tracking_period_worksheet.write(counter, 13, "", money_blue_cell) # not printed for workpackages
                    tracking_period_worksheet.write(counter, 14, "", money_blue_cell) # not printed for workpackages
                    tracking_period_worksheet.write(counter, 15, "", blue_cell) # NECESSARY EMPTY field for workpackages!
                    tracking_period_worksheet.write(counter, 16, "", money_blue_cell) # not printed for workpackages
                    tracking_period_worksheet.write(counter, 17, "", money_blue_cell) # not printed for workpackages

                    tracking_period_worksheet.write_number(counter, 18, atr.actual_cost, money_blue_cell)
                    tracking_period_worksheet.write(counter, 19, "", money_blue_cell) # not printed for workpackages
                    #percentage_completed = str(atr.percentage_completed) + "%"
                    tracking_period_worksheet.write(counter, 20, atr.percentage_completed / 100.0, percent_blue_cell)
                    tracking_period_worksheet.write(counter, 21, "", blue_cell) # not printed for workpackages
                    tracking_period_worksheet.write_number(counter, 22, atr.earned_value, money_blue_cell)
                    tracking_period_worksheet.write_number(counter, 23, atr.planned_value, money_blue_cell)
                else:
                    # activity:
                    tracking_period_worksheet.write_number(counter, 0, atr.activity.activity_id, green_cell)
                    tracking_period_worksheet.write(counter, 1, atr.activity.name, green_cell)
                    tracking_period_worksheet.write_datetime(counter, 2, atr.activity.baseline_schedule.start, date_gray_cell)
                    tracking_period_worksheet.write_datetime(counter, 3, atr.activity.baseline_schedule.end, date_gray_cell)
                    tracking_period_worksheet.write(counter, 4, self.get_duration_str(atr.activity.baseline_schedule.duration), gray_cell)
                    self.write_resources(tracking_period_worksheet, counter, 5, atr.activity.resources, gray_cell)
                    tracking_period_worksheet.write_number(counter, 6, atr.activity.resource_cost, money_gray_cell)
                    tracking_period_worksheet.write_number(counter, 7, atr.activity.baseline_schedule.fixed_cost, money_gray_cell)
                    tracking_period_worksheet.write_number(counter, 8, atr.activity.baseline_schedule.hourly_cost, money_gray_cell)
                    tracking_period_worksheet.write_number(counter, 9, atr.activity.baseline_schedule.var_cost, money_gray_cell)
                    tracking_period_worksheet.write_number(counter, 10, atr.activity.baseline_schedule.total_cost, money_gray_cell)
                    if atr.actual_start and atr.actual_start.date() < datetime.datetime.max.date():
                        tracking_period_worksheet.write_datetime(counter, 11, atr.actual_start, date_green_cell)
                    else:
                        tracking_period_worksheet.write(counter, 11, '', green_cell)

                    tracking_period_worksheet.write(counter, 12, self.get_duration_str(atr.actual_duration), green_cell)
                    tracking_period_worksheet.write_number(counter, 13, atr.planned_actual_cost, money_gray_cell)
                    tracking_period_worksheet.write_number(counter, 14, atr.planned_remaining_cost, money_gray_cell)
                    tracking_period_worksheet.write(counter, 15, self.get_duration_str(atr.remaining_duration), gray_cell)
                    tracking_period_worksheet.write_number(counter, 16, atr.deviation_pac, money_gray_cell)
                    tracking_period_worksheet.write_number(counter, 17, atr.deviation_prc, money_yellow_cell)
                    tracking_period_worksheet.write_number(counter, 18, atr.actual_cost, money_green_cell)
                    tracking_period_worksheet.write_number(counter, 19, atr.remaining_cost, money_gray_cell)
                    tracking_period_worksheet.write(counter, 20, atr.percentage_completed / 100.0, percent_green_cell)
                    tracking_period_worksheet.write(counter, 21, atr.tracking_status, gray_cell)
                    tracking_period_worksheet.write_number(counter, 22, atr.earned_value, money_gray_cell)
                    tracking_period_worksheet.write_number(counter, 23, atr.planned_value, money_gray_cell)
                counter += 1

        # Write the agenda
        agenda_worksheet = workbook.add_worksheet("Agenda")
        agenda_worksheet.set_column(0, 0, 12)
        agenda_worksheet.set_column(3, 3, 8)
        agenda_worksheet.set_column(6, 6, 12)
        agenda_worksheet.merge_range('A1:B1', 'Working Hours', header)
        agenda_worksheet.merge_range('D1:E1', 'Working Days', header)
        agenda_worksheet.write('G1', 'Holidays (Optional)', header)
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for i in range(0, 24):
            hour_string = str(i) + ":00 - " + str((i+1)%24) + ":00"
            agenda_worksheet.write(i+1, 0, hour_string, gray_cell)
            if project_object.agenda.working_hours[i]:
                agenda_worksheet.write(i+1, 1, "Yes", green_cell_agenda)
            else:
                agenda_worksheet.write(i+1, 1, "No", red_cell_agenda)
        for i in range(0, 7):
            agenda_worksheet.write(i+1, 3, days[i], gray_cell)
            if project_object.agenda.working_days[i]:
                agenda_worksheet.write(i+1, 4, "Yes", green_cell_agenda)
            else:
                agenda_worksheet.write(i+1, 4, "No", red_cell_agenda)
        agenda_worksheet.conditional_format('B2:B25', { 'type':     'cell',
                                                        'criteria': '==',
                                                        'value':    '"Yes"',
                                                        'format':   green_cell_agenda})
        agenda_worksheet.conditional_format('B2:B25', { 'type':     'cell',
                                                        'criteria': '==',
                                                        'value':    '"No"',
                                                        'format':   red_cell_agenda})
        agenda_worksheet.conditional_format('E2:E8', { 'type':     'cell',
                                                        'criteria': '==',
                                                        'value':    '"Yes"',
                                                        'format':   green_cell_agenda})
        agenda_worksheet.conditional_format('E2:E8', { 'type':     'cell',
                                                        'criteria': '==',
                                                        'value':    '"No"',
                                                        'format':   red_cell_agenda})
        counter = 1
        for holiday in project_object.agenda.holidays:
            agenda_worksheet.write(counter, 6, holiday, holiday_yellow_cell)
            counter += 1

        # Write the tracking overview
        overview_worksheet = workbook.add_worksheet("Tracking Overview")
        overview_worksheet.set_column(0, 13, 15)
        overview_worksheet.set_column(14, 30, 15)
        overview_worksheet.set_column(14, 30, 17)
        overview_worksheet.set_row(1, 30)

        overview_worksheet.merge_range('A1:C1', 'General', header)
        overview_worksheet.merge_range('D1:G1', 'EVM Performance Measures', header)
        overview_worksheet.merge_range('H1:N1', 'EVM Forecasting', header)

        overview_worksheet.write('A2', "Name", header)
        overview_worksheet.write('B2', "Start Tracking Period", header)
        overview_worksheet.write('C2', "Status date", header)
        overview_worksheet.write('D2', "Planned Value (PV)", header)
        overview_worksheet.write('E2', "Earned Value (EV)", header)
        overview_worksheet.write('F2', "Actual Cost (AC)", header)
        overview_worksheet.write('G2', "Earned Schedule (ES)", header)
        overview_worksheet.write('H2', "Schedule Variance (SV)", header)
        overview_worksheet.write('I2', "Schedule Performance Index (SPI)", header)
        overview_worksheet.write('J2', "Cost Variance (CV)", header)
        overview_worksheet.write('K2', "Cost Performance Index (CPI)", header)
        overview_worksheet.write('L2', "Schedule Variance (SV(t))", header)
        overview_worksheet.write('M2', "Schedule Performance Index (SPI(t))", header)
        overview_worksheet.write('N2', "p-factor", header)
        overview_worksheet.write('O2', "EAC(t)-PV (PF=1)", header)
        overview_worksheet.write('P2', "EAC(t)-PV (PF=SPI)", header)
        overview_worksheet.write('Q2', "EAC(t)-PV (PF=SCI)", header)
        overview_worksheet.write('R2', "EAC(t)-ED (PF=1)", header)
        overview_worksheet.write('S2', "EAC(t)-ED (PF=SPI)", header)
        overview_worksheet.write('T2', "EAC(t)-ED (PF=SPI)", header)
        overview_worksheet.write('U2', "EAC(t)-ES (PF=1)", header)
        overview_worksheet.write('V2', "EAC(t)-ES (PF=SPI(t))", header)
        overview_worksheet.write('W2', "EAC(t)-ES (PF=SCI(t))", header)
        overview_worksheet.write('X2', "EAC (PF=1)", header)
        overview_worksheet.write('Y2', "EAC (PF=CPI)", header)
        overview_worksheet.write('Z2', "EAC (PF=SPI)", header)
        overview_worksheet.write('AA2', "EAC (PF=SPI(t))", header)
        overview_worksheet.write('AB2', "EAC (PF=SCI)", header)
        overview_worksheet.write('AC2', "EAC (PF=SCI(t))", header)
        overview_worksheet.write('AD2', "EAC (PF=0.8*CPI+0.2*SPI)", header)
        overview_worksheet.write('AE2', "EAC (PF=0.8*CPI+0.2*SPI(t))", header)

        # generate PV curve:
        generatedPVcurve = self.calculate_PVcurve(project_object)
        # DEBUG
        #workbookFilepath, fileextension = os.path.splitext(file_path_output)
        #with open(workbookFilepath + "-PV.csv", "w", newline='') as csvfile:
        #    PVwriter = csv.writer(csvfile, delimiter=';')
        #    PVwriter.writerow(["PV(t)", "t"])
        #    for PVrow in generatedPVcurve:
        #        PVwriter.writerow([PVrow[0], PVrow[1].strftime("%d/%m/%Y %H:%M:%S")])

        counter = 2
        for tracking_period in project_object.tracking_periods:
            overview_worksheet.write(counter, 0, tracking_period.tracking_period_name, gray_cell)
            index = project_object.tracking_periods.index(tracking_period)
            if index == 0:
                overview_worksheet.write_datetime(counter, 1, min([atr.activity.baseline_schedule.start for atr in tracking_period.tracking_period_records]), date_gray_cell)
            else:
                overview_worksheet.write_datetime(counter, 1, project_object.tracking_periods[index-1].tracking_period_statusdate, date_gray_cell)
            overview_worksheet.write_datetime(counter, 2, tracking_period.tracking_period_statusdate, date_gray_cell)

            # calculate the key metrics and write them out
            PV, EV, AC = self.calculate_aggregated_keyMetrics(tracking_period)
            overview_worksheet.write_number(counter, 3, PV, money_gray_cell)
            overview_worksheet.write_number(counter, 4, EV, money_gray_cell)
            overview_worksheet.write_number(counter, 5, AC, money_gray_cell)
            # calculate ES
            ES = self.calculate_es(project_object, generatedPVcurve, EV, PV, tracking_period.tracking_period_statusdate)
            overview_worksheet.write_datetime(counter, 6, ES, date_gray_cell)
            # calculate SV
            sv = EV - PV
            overview_worksheet.write_number(counter, 7, sv, money_gray_cell)
            # calculate SPI
            if not PV:
                spi = 0.0
            else:
                spi = EV / PV
            # save spi value also in tracking_period for visualisations:
            tracking_period.spi = spi
            overview_worksheet.write(counter, 8, spi, percent_gray_cell)

            # calculate CV
            cv = EV - AC
            overview_worksheet.write_number(counter, 9, cv, money_gray_cell)
            if not AC:
                cpi = 0.0
            else:
                cpi = EV /AC
            # save cpi value also in tracking_period for visualisations:
            tracking_period.cpi = cpi
            overview_worksheet.write(counter, 10, cpi, percent_gray_cell)

            # calculate SV(t)
            sv_t, sv_t_str = self.calculate_SVt(project_object, ES, tracking_period.tracking_period_statusdate)
            # save SV(t) value also in tracking_period for visualisations:
            tracking_period.sv_t = sv_t
            overview_worksheet.write(counter, 11, sv_t_str, gray_cell)

            # calculate SPI(t)
            spi_t = self.calculate_SPIt(project_object, ES, tracking_period.tracking_period_statusdate)
            # save spi_t value also in tracking_period for visualisations:
            tracking_period.spi_t = spi_t
            overview_worksheet.write(counter, 12, spi_t, percent_gray_cell)

            # calculate p-factor
            p_factor = self.calculate_p_factor(project_object, tracking_period, ES)
            overview_worksheet.write(counter, 13, p_factor, percent_gray_cell)
            # save p_factor value also in tracking_period for visualisations:
            tracking_period.p_factor = p_factor

            workingHours_inDay = project_object.agenda.get_working_hours_in_a_day()
            BAC = generatedPVcurve[-1][0]  # last PV cumsum point corresponds to BAC
            PD = project_object.agenda.get_time_between(generatedPVcurve[0][1], generatedPVcurve[-1:][0][1])
            PD_workingHours = PD.days * workingHours_inDay + round(PD.seconds / 3600)  # represent Project Duration in workinghours
            AT_duration = project_object.agenda.get_time_between(generatedPVcurve[0][1], tracking_period.tracking_period_statusdate)
            AT_duration_workingHours = AT_duration.days * workingHours_inDay + round(AT_duration.seconds / 3600)  # represent AT duration since project start in workinghours
            ES_duration = project_object.agenda.get_time_between(generatedPVcurve[0][1], ES)
            ES_duration_workingHours = ES_duration.days * workingHours_inDay + round(ES_duration.seconds / 3600)  # represents the ES duration since project start in workinghours

            # write EAC(t) - Planned Value method (PF = 1)
            PVrate = BAC / float(PD_workingHours) if bool(PD_workingHours) else 0
            TV = sv / PVrate if bool(PVrate) else 0  # Time variance
            EAC_t_pv1 = PD_workingHours - TV
            EAC_t_pv1_days = int(EAC_t_pv1 / workingHours_inDay)
            overview_worksheet.write_datetime(counter, 14, project_object.agenda.get_end_date(generatedPVcurve[0][1], EAC_t_pv1_days, round(EAC_t_pv1 - EAC_t_pv1_days * workingHours_inDay)), date_gray_cell)

            # write EAC(t) - Planned Value method (PF = spi)
            EAC_t_pv2 = PD_workingHours / spi if bool(spi)  else 0
            EAC_t_pv2_days = int(EAC_t_pv2 / workingHours_inDay)
            overview_worksheet.write_datetime(counter, 15, project_object.agenda.get_end_date(generatedPVcurve[0][1], EAC_t_pv2_days, round(EAC_t_pv2 - EAC_t_pv2_days * workingHours_inDay)), date_gray_cell)

            # write EAC(t) - Planned Value method (PF = SCI = SPI * CPI)
            EAC_t_pv3 = PD_workingHours / (spi * cpi) if bool(spi * cpi)  else 0
            EAC_t_pv3_days = int(EAC_t_pv3 / workingHours_inDay)
            overview_worksheet.write_datetime(counter, 16, project_object.agenda.get_end_date(generatedPVcurve[0][1], EAC_t_pv3_days, round(EAC_t_pv3 - EAC_t_pv3_days * workingHours_inDay)), date_gray_cell)

            # write EAC(t) - Earned duration method (PF = 1)
            ED = AT_duration_workingHours * spi
            EAC_t_ed1 = AT_duration_workingHours + (max(PD_workingHours, AT_duration_workingHours) - ED)
            EAC_t_ed1_days = int(EAC_t_ed1 / workingHours_inDay)
            overview_worksheet.write_datetime(counter, 17, project_object.agenda.get_end_date(generatedPVcurve[0][1], EAC_t_ed1_days, round(EAC_t_ed1 - EAC_t_ed1_days * workingHours_inDay)), date_gray_cell)

            # write EAC(t) - Earned duration method (PF = spi)
            if bool(spi):
                EAC_t_ed2 = AT_duration_workingHours + (max(PD_workingHours, AT_duration_workingHours) - ED) / spi
            else:
                EAC_t_ed2 = AT_duration_workingHours
            EAC_t_ed2_days = int(EAC_t_ed2 / workingHours_inDay)
            overview_worksheet.write_datetime(counter, 18, project_object.agenda.get_end_date(generatedPVcurve[0][1], EAC_t_ed2_days, round(EAC_t_ed2 - EAC_t_ed2_days * workingHours_inDay)), date_gray_cell)

            # write EAC(t) - Earned duration method (PF = SCI = SPI * CPI)
            if bool(spi) and bool(cpi) :
                EAC_t_ed3 = AT_duration_workingHours + (max(PD_workingHours, AT_duration_workingHours) - ED) / (spi * cpi)
            else:
                EAC_t_ed3 = AT_duration_workingHours
            EAC_t_ed3_days = int(EAC_t_ed3 / workingHours_inDay)
            overview_worksheet.write_datetime(counter, 19, project_object.agenda.get_end_date(generatedPVcurve[0][1], EAC_t_ed3_days, round(EAC_t_ed3 - EAC_t_ed3_days * workingHours_inDay)), date_gray_cell)

            # write EAC(t) - Earned schedule method (PF = 1)
            EAC_t_es1 = AT_duration_workingHours + (PD_workingHours - ES_duration_workingHours)
            EAC_t_es1_days = int(EAC_t_es1 / workingHours_inDay)
            overview_worksheet.write_datetime(counter, 20, project_object.agenda.get_end_date(generatedPVcurve[0][1], EAC_t_es1_days, round(EAC_t_es1 - EAC_t_es1_days * workingHours_inDay)), date_gray_cell)

            # write EAC(t) - Earned schedule method (PF = spi_t)
            if bool(spi_t):
                EAC_t_es2 = AT_duration_workingHours + (PD_workingHours - ES_duration_workingHours) / spi_t
            else:
                EAC_t_es2 = AT_duration_workingHours
            EAC_t_es2_days = int(EAC_t_es2 / workingHours_inDay)
            overview_worksheet.write_datetime(counter, 21, project_object.agenda.get_end_date(generatedPVcurve[0][1], EAC_t_es2_days, round(EAC_t_es2 - EAC_t_es2_days * workingHours_inDay)), date_gray_cell)

            # write EAC(t) - Earned schedule method (PF = SCI = SPI(t) * CPI)
            if bool(spi_t) and bool(cpi):
                EAC_t_es3 = AT_duration_workingHours + (PD_workingHours - ES_duration_workingHours) / (spi_t * cpi)
            else:
                EAC_t_es3 = AT_duration_workingHours
            EAC_t_es3_days = int(EAC_t_es3 / workingHours_inDay)
            overview_worksheet.write_datetime(counter, 22, project_object.agenda.get_end_date(generatedPVcurve[0][1], EAC_t_es3_days, round(EAC_t_es3 - EAC_t_es3_days * workingHours_inDay)), date_gray_cell)


            # write EAC(PF = 1)
            overview_worksheet.write_number(counter, 23, self.calculate_eac(AC, BAC, EV, 1), money_gray_cell)
            # write EAC(PF = cpi)
            overview_worksheet.write_number(counter, 24, self.calculate_eac(AC,BAC, EV, cpi), money_gray_cell)

            # write EAC(PF = spi)
            overview_worksheet.write_number(counter, 25, self.calculate_eac(AC, BAC, EV, spi), money_gray_cell)

            # write EAC(PF = spi_t)
            overview_worksheet.write_number(counter, 26, self.calculate_eac(AC, BAC, EV, spi_t), money_gray_cell)

            # write EAC(PF = SCI = SPI * CPI)
            overview_worksheet.write_number(counter, 27, self.calculate_eac(AC, BAC, EV, spi * cpi), money_gray_cell)

            # write EAC(PF = SCI(t) = SPI(t) * CPI
            overview_worksheet.write_number(counter, 28, self.calculate_eac(AC, BAC, EV, spi_t * cpi), money_gray_cell)

            # write EAC(PF = 0.8*CPI+0.2*SPI)
            overview_worksheet.write_number(counter, 29, self.calculate_eac(AC, BAC, EV, 0.8*cpi+0.2*spi), money_gray_cell)

            # write EAC(PF = 0.8*CPI+0.2*SPI(t))
            overview_worksheet.write_number(counter, 30, self.calculate_eac(AC, BAC, EV, 0.8*cpi+0.2*spi_t), money_gray_cell)


            counter += 1

        return workbook

    @staticmethod
    def get_duration_str(delta, negativeValue=False):
        if delta:
            # Writing a duration requires some converting..
            if delta.days != 0 and delta.seconds != 0:
                if not negativeValue:
                    duration = str(delta.days) + "d " + str(round(delta.seconds / 3600.0)) + "h"
                else:
                    duration = "-" + str(delta.days) + "d " + str(round(delta.seconds / 3600.0)) + "h"
            elif delta.seconds != 0:
                if not negativeValue:
                    duration = str(round(delta.seconds / 3600.0)) + "h"
                else:
                    duration = "-" + str(round(delta.seconds / 3600.0)) + "h"
            else:
                if not negativeValue:
                    duration = str(delta.days) + "d"
                else:
                    duration = "-" + str(delta.days) + "d"
            return duration
        return "0"

    @staticmethod
    def get_duration_hours_str(delta, workingHoursInDay):
        if delta:
            totalWorkingHours = delta.days * workingHoursInDay + int(delta.seconds / 3600)
            return "{0}h".format(totalWorkingHours)
        else:
            return "0"

    @staticmethod
    def write_wbs(worksheet, row, column, wbs, format):
        # Instead of writing (x, y, z) write x.y.z
        to_write = ''
        if len(wbs) > 0:
            for i in range(0, len(wbs)-1):
                to_write += str(wbs[i]) + '.'
            to_write += str(wbs[-1])
        worksheet.write(row, column, to_write, format)

    @staticmethod
    def write_successors(worksheet, row, column, successors, format, agenda):
        # Write in format of FSxx;SSxx..
        to_write = ''
        if len(successors) > 0:
            for i in range(0, len(successors)-1):
                to_write += str(successors[i][1]) + str(successors[i][0])
                if successors[i][2] != 0:
                    to_write += ("+" if successors[i][2] > 0 else "-") + agenda.convert_workingHours_to_durationString(abs(successors[i][2])) # convert workinghours to formatted string
                to_write += ";"
            to_write += str(successors[-1][1]) + str(successors[-1][0])
            if successors[-1][2] != 0:
                to_write += ("+" if successors[-1][2] > 0 else "-") + agenda.convert_workingHours_to_durationString(abs(successors[-1][2])) # convert workinghours to formatted string
        worksheet.write(row, column, to_write, format)

    @staticmethod
    def write_predecessors(worksheet, row, column, predecessors, format, agenda):
        # Write in format of xxSF;xxFF..
        to_write = ''
        if len(predecessors) > 0:
            for i in range(0, len(predecessors)-1):
                to_write += str(predecessors[i][0]) + str(predecessors[i][1])
                if predecessors[i][2] != 0:
                    to_write += ("+" if predecessors[i][2] > 0 else "-") + agenda.convert_workingHours_to_durationString(abs(predecessors[i][2])) # convert workinghours to formatted string
                to_write += ";"
            to_write += str(predecessors[-1][0]) + str(predecessors[-1][1])
            if predecessors[-1][2] != 0:
                to_write += ("+" if predecessors[-1][2] > 0 else "-") + agenda.convert_workingHours_to_durationString(abs(predecessors[-1][2])) # convert workinghours to formatted string
        worksheet.write(row, column, to_write, format)

    @staticmethod
    def write_resources(worksheet, row, column, resources, format):
        # Write out the resources in a specific format
        to_write = ''
        if len(resources) > 0:
            for i in range(0, len(resources)-1):
                to_write += resources[i][0].name
                if resources[i][1] != 1 or resources[i][2]:
                    # if not a unit 1 demand or if fixed assignment:
                    to_write += "[" +  "{0:.2f}".format(float(resources[i][1]))
                    if resources[i][2]: to_write += "F" # fixed resource assignment
                    to_write += " #" + str(resources[i][0].availability) + "];"
                else:
                    to_write += ";"
            to_write += resources[-1][0].name
            if resources[-1][1] != 1 or resources[-1][2]:
                # if not a unit 1 demand or if fixed assignment:
                to_write += "[" + "{0:.2f}".format(float(resources[-1][1]))
                if resources[-1][2]: to_write += "F" # fixed resource assignment
                to_write += " #" + str(resources[-1][0].availability) + "]"
        worksheet.write(row, column, to_write, format)

    @staticmethod
    def write_resource_assign_cost(worksheet, row, column, resource, activities, format, moneyformat, project_object):
        # For every resource, we check to which activities it is assigned and what the total cost is
        to_write = ''
        cost = 0
        for activity in activities:
            for _resource in activity.resources:
                if resource == _resource[0]:
                    if _resource[1] == 1 and not _resource[2]:
                        to_write += str(activity.activity_id) + ';'
                    else:
                        to_write += str(activity.activity_id) + '[' + "{0:.2f}".format(float(_resource[1]))
                        if _resource[2]: to_write += "F" # fixed resource assignment
                        to_write += ' ' + str(resource.name) + '];' #TODO: this used to be resource.availability but needs to be somth else

                    ## calculate cost of the given resource assignment and add it to the total resource cost: # is present in project_object
                    #cost += (activity.baseline_schedule.duration.days*project_object.agenda.get_working_hours_in_a_day() +
                    #         activity.baseline_schedule.duration.seconds/3600)*(_resource[1]*resource.cost_unit)
        worksheet.write(row, column, to_write, format)
        worksheet.write_number(row, column+1, resource.total_resource_cost, moneyformat)

